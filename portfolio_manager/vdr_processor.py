"""
VDR Document Processor
======================
Processes uploaded documents for site data extraction.

Supports:
- PDF (PyPDF2)
- Word (python-docx)
- Excel (openpyxl)
"""

import io
import json
import streamlit as st
import google.generativeai as genai
from typing import Dict, Any, Optional

try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def extract_text_from_pdf(uploaded_file) -> str:
    """Extract text from PDF file."""
    if not PDF_AVAILABLE:
        raise ImportError("PyPDF2 not installed")
    
    try:
        pdf_reader = PyPDF2.PdfReader(uploaded_file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        raise Exception(f"PDF extraction failed: {str(e)}")


def extract_text_from_docx(uploaded_file) -> str:
    """Extract text from Word document."""
    if not DOCX_AVAILABLE:
        raise ImportError("python-docx not installed")
    
    try:
        doc = Document(uploaded_file)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        raise Exception(f"DOCX extraction failed: {str(e)}")


def extract_text_from_xlsx(uploaded_file) -> str:
    """Extract text from Excel file."""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl not installed")
    
    try:
        wb = openpyxl.load_workbook(uploaded_file)
        text = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"\n--- Sheet: {sheet_name} ---\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                text += row_text + "\n"
        return text
    except Exception as e:
        raise Exception(f"Excel extraction failed: {str(e)}")


def extract_text_from_csv_txt(uploaded_file) -> str:
    """Extract text from CSV or TXT file."""
    try:
        # Read as text
        text = uploaded_file.read().decode('utf-8', errors='ignore')
        return text
    except Exception as e:
        raise Exception(f"Text file extraction failed: {str(e)}")


def process_uploaded_file(uploaded_file) -> str:
    """Process uploaded file and extract text."""
    filename = uploaded_file.name
    ext = filename.split('.')[-1].lower()
    
    if ext == 'pdf':
        return extract_text_from_pdf(uploaded_file)
    elif ext == 'docx':
        return extract_text_from_docx(uploaded_file)
    elif ext in ['xlsx', 'xls']:
        return extract_text_from_xlsx(uploaded_file)
    elif ext in ['txt', 'csv']:
        return extract_text_from_csv_txt(uploaded_file)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def extract_site_data_from_text(text: str, filename: str) -> Optional[Dict[str, Any]]:
    """Use Gemini LLM to extract structured site data from document text."""
    
    # Limit text to avoid token limits (keep first 20k chars)
    text_sample = text[:20000]
    
    extraction_prompt = f"""Analyze this document and extract data center site information.

Document Name: {filename}
Content:
{text_sample}

Extract the following fields if mentioned (return null if not found):

{{
  "site_name": "Site name or project name",
  "state": "2-letter state code (TX, OK, GA, etc.)",
  "utility": "Utility name (Oncor, PSO, AEP, Duke Energy, Georgia Power, Dominion, etc.)",
  "target_mw": "Total interconnection MW capacity",
  "acreage": "Integer acreage",
  "study_status": "One of: not_started, screening_study, contract_study, loa, energy_contract",
  "land_control": "One of: owned, option, loi, negotiating, none",
  "voltage": "Interconnection voltage (e.g., 138kV, 345kV)",
  "iso": "ISO/RTO (PJM, ERCOT, SPP, MISO, etc.)",
  "substation": "Nearest substation name",
  "transmission_distance": "Distance to transmission line in miles",
  "service_type": "Network or Radial service",
  "interconnection_cost": "Estimated interconnection cost in millions",
  "developer": "Developer name if mentioned",
  "queue_position": "Queue position number if mentioned",
  
  "timeline_to_cod": "Timeline description (e.g., '36 months from agreement execution', '24 months from approval')",
  "cod_date": "Commercial Operation Date in YYYY-MM-DD if explicitly stated",
  "agreement_execution_date": "Date when the final interconnection/facilities agreement was executed (look for: IA execution, FA execution, Interconnection Agreement, Facilities Agreement, GIA, LGIA date)",
  
  "phases": [
    {{
      "phase_number": "Phase number (1, 2, etc.)",
      "interconnection_mw": "Interconnection capacity for this phase in MW",
      "cod_date": "COD for this phase in YYYY-MM-DD",
      "timeline": "Timeline description for this phase"
    }}
  ],
  
  "generation": {{
    "gas_mw": "Natural gas generation capacity in MW",
    "solar_mw": "Solar generation capacity in MW",
    "battery_mw": "Battery storage capacity in MW",
    "battery_mwh": "Battery storage energy capacity in MWh",
    "other_mw": "Other generation types in MW"
  }},
  
  "load_profile": "Description of expected load growth or ramp (e.g., '200MW per year starting 2025')",
  "notes": "Any other relevant information including POI details, timeline notes, risks, opportunities"
}}

Study status mappings (map old terminology to new phasing structure):
- "System Impact Study" / "SIS" / "Screening Study" → screening_study
- "Facilities Study" / "FS" / "Contract Study" → contract_study
- "Facilities Agreement" / "FA" / "Letter of Agreement" / "LOA" → loa
- "Interconnection Agreement" / "IA" / "Energy Contract" → energy_contract

Timeline parsing instructions:
- If you see "X months from [event]", extract timeline_to_cod and the reference event
- Calculate actual dates when possible based on today's date (December 1, 2025)
- Extract phase-specific timelines if project is described in multiple phases

Return ONLY valid JSON. Use null for unknown values."""

    try:
        api_key = st.secrets.get("GEMINI_API_KEY")
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('models/gemini-2.0-flash-exp')
        
        response = model.generate_content(
            extraction_prompt,
            generation_config={"response_mime_type": "application/json"}
        )
        
        extracted = json.loads(response.text)
        
        # Handle list responses
        if isinstance(extracted, list):
            extracted = extracted[0] if extracted else {}
        
        # Clean null values
        extracted = {k: v for k, v in extracted.items() if v is not None and v != "null"}
        
        # Add metadata
        extracted['_source_file'] = filename
        
        return extracted if extracted else None
        
    except Exception as e:
        st.error(f"LLM extraction failed: {str(e)}")
        return None


def upload_to_google_drive(file_bytes: bytes, filename: str, folder_id: str) -> Optional[str]:
    """Upload file to Google Drive folder."""
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaInMemoryUpload
        from google.oauth2.service_account import Credentials
        
        # Get credentials from Streamlit secrets
        creds_dict = dict(st.secrets["gcp_service_account"])
        scopes = ['https://www.googleapis.com/auth/drive.file']
        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        
        # Build Drive service
        service = build('drive', 'v3', credentials=credentials)
        
        # File metadata
        file_metadata = {
            'name': filename,
            'parents': [folder_id]
        }
        
        # Upload file
        media = MediaInMemoryUpload(file_bytes, mimetype='application/octet-stream', resumable=True)
        
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id, webViewLink'
        ).execute()
        
        return file.get('webViewLink')
        
    except Exception as e:
        # Don't fail the whole process if Drive upload fails
        st.warning(f"⚠️ Google Drive upload failed: {str(e)}")
        st.info("Data extraction will continue without Drive upload. Make sure the VDR Uploads folder is shared with your service account.")
        return None
